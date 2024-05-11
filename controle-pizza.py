import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime

def update_stock(flavor, quantity, operation):
    # Carregar ou criar o arquivo Excel
    try:
        wb = load_workbook('controle_pizza.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Date', 'Operation', 'Flavor', 'Quantity', 'Stock'])

    # Atualizar o estoque
    current_stock = 0
    for row in sheet.iter_rows(min_row=2, max_col=4):
        if row[2].value.lower() == flavor:
            if operation == 'Entrada':
                current_stock += row[3].value
            elif operation == 'Saída':
                current_stock -= row[3].value  # Subtraído o valor da saída do estoque

    new_stock = current_stock + quantity if operation == 'Entrada' else current_stock - quantity
    sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), operation, flavor, quantity, new_stock])

    # Salvar o arquivo Excel
    wb.save('controle_pizza.xlsx')

def register_entry():
    flavor = entry_flavor.get().strip().lower()
    if flavor not in ['presunto e queijo', 'calabresa']:
        messagebox.showerror('Erro', 'Sabor inválido. Por favor, escolha entre "Presunto e queijo" ou "Calabresa".')
        return
    quantity = int(entry_quantity.get())
    update_stock(flavor, quantity, 'Entrada')
    messagebox.showinfo('Sucesso', f'{quantity} mini pizzas de {flavor} registradas como entrada no estoque.')

def register_exit():
    flavor = entry_flavor.get().strip().lower()
    if flavor not in ['presunto e queijo', 'calabresa']:
        messagebox.showerror('Erro', 'Sabor inválido. Por favor, escolha entre "Presunto e queijo" ou "Calabresa".')
        return
    quantity = int(entry_quantity.get())
    update_stock(flavor, quantity, 'Saída')
    messagebox.showinfo('Sucesso', f'{quantity} mini pizzas de {flavor} registradas como saída do estoque.')

def clear_stock():
    confirmation = messagebox.askyesno('Confirmação', 'Tem certeza de que deseja limpar todo o estoque?')
    if confirmation:
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Date', 'Operation', 'Flavor', 'Quantity', 'Stock'])
        wb.save('controle_pizza.xlsx')
        messagebox.showinfo('Sucesso', 'Todo o estoque foi apagado.')


def check_stock():
    try:
        wb = load_workbook('controle_pizza.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        messagebox.showerror('Erro', 'O arquivo do controle de estoque não foi encontrado.')
        return

    # Dicionários para armazenar o estoque de cada sabor de pizza
    stock_dict_presunto_queijo = {'presunto e queijo': 0}
    stock_dict_calabresa = {'calabresa': 0}

    # Percorrendo todas as linhas para calcular o estoque de cada sabor de pizza
    for row in sheet.iter_rows(min_row=2, max_col=5):
        flavor = row[2].value.strip().lower()
        quantity = row[3].value
        operation = row[1].value

        if flavor == 'presunto e queijo':
            if operation == 'Entrada':
                stock_dict_presunto_queijo[flavor] += quantity
            elif operation == 'Saída':
                stock_dict_presunto_queijo[flavor] -= quantity
        elif flavor == 'calabresa':
            if operation == 'Entrada':
                stock_dict_calabresa[flavor] += quantity
            elif operation == 'Saída':
                stock_dict_calabresa[flavor] -= quantity

    # Calculando o total de cada sabor de pizza
    total_presunto_queijo = stock_dict_presunto_queijo['presunto e queijo']
    total_calabresa = stock_dict_calabresa['calabresa']

    # Calculando o total geral
    total_geral = total_presunto_queijo + total_calabresa

    # Exibindo o total de estoque para cada sabor de pizza e o total geral
    messagebox.showinfo('Estoque Atual', f'Total de mini pizzas de presunto e queijo: {total_presunto_queijo}\nTotal de mini pizzas de calabresa: {total_calabresa}\nTotal geral: {total_geral}.')



def remove_item():
    flavor = entry_flavor.get().strip().lower()
    confirmation = messagebox.askyesno('Confirmação', f'Tem certeza de que deseja apagar todas as entradas/saídas de {flavor}?')
    if confirmation:
        wb = load_workbook('controle_pizza.xlsx')
        sheet = wb.active
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2, max_col=4):
            if row[2].value.lower() == flavor:
                rows_to_delete.append(row)
        for row in rows_to_delete:
            sheet.delete_rows(row[0].row)
        wb.save('controle_pizza.xlsx')
        messagebox.showinfo('Sucesso', f'Todas as entradas/saídas de {flavor} foram apagadas.')

# Configuração da janela principal
root = tk.Tk()
root.title('Contole de Pizzas Julinho!')

# Ícone de pizza
pizza_icon = tk.PhotoImage(file='1404945.png')
root.iconphoto(True, pizza_icon)

# Cores e fontes
bg_color = "#FFD700"  # Amarelo dourado
button_color = "#CD5C5C"  # Vermelho escuro
font_family = "Arial"
font_size = 12

root.configure(bg=bg_color)

# Componentes da interface
label_flavor = tk.Label(root, text='Sabor da Pizza:', bg=bg_color, font=(font_family, font_size))
label_flavor.grid(row=0, column=0, padx=20, pady=10)

entry_flavor = tk.Entry(root, width=30, font=(font_family, font_size))
entry_flavor.grid(row=0, column=1, padx=20, pady=10)

label_quantity = tk.Label(root, text='Quantidade:', bg=bg_color, font=(font_family, font_size))
label_quantity.grid(row=1, column=0, padx=20, pady=10)

entry_quantity = tk.Entry(root, width=30, font=(font_family, font_size))
entry_quantity.grid(row=1, column=1, padx=20, pady=10)

button_entry = tk.Button(root, text='Registrar Entrada', command=register_entry, width=20, bg=button_color, font=(font_family, font_size))
button_entry.grid(row=2, column=0, padx=20, pady=10)

button_exit = tk.Button(root, text='Registrar Saída', command=register_exit, width=20, bg=button_color, font=(font_family, font_size))
button_exit.grid(row=2, column=1, padx=20, pady=10)

button_clear = tk.Button(root, text='Limpar Estoque', command=clear_stock, width=20, bg=button_color, font=(font_family, font_size))
button_clear.grid(row=3, column=0, padx=20, pady=10)

button_remove = tk.Button(root, text='Remover Item', command=remove_item, width=20, bg=button_color, font=(font_family, font_size))
button_remove.grid(row=3, column=1, padx=20, pady=10)

button_check_stock = tk.Button(root, text='Verificar Estoque', command=check_stock, width=40, bg=button_color, font=(font_family, font_size))
button_check_stock.grid(row=4, column=0, columnspan=2, padx=20, pady=10)
# Loop principal da interface
root.mainloop()
